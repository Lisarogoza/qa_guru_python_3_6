import os
from os.path import basename
from zipfile import ZipFile
from PyPDF2 import PdfReader
from openpyxl import load_workbook

files_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files')
zip_path = os.path.join(files_path, "archived.zip")


def test_archive_create():
    file_dir = os.listdir(files_path)
    with ZipFile(zip_path, "w") as zip_:
        for file in file_dir:
            add_file = os.path.join(files_path, file)
            zip_.write(add_file, basename(add_file))
    assert zip_.namelist() == ['file_example_XLSX_50.xlsx', 'Lapin_giftcard.pdf', 'users.csv']


def test_xlsx():
    with ZipFile(zip_path) as arch:
        arch.extract('file_example_XLSX_50.xlsx')
        workbook = load_workbook('file_example_XLSX_50.xlsx')
        sheet = workbook.active
        check_value = str(sheet.cell(row=3, column=2).value)
        assert check_value == 'Mara'
        os.remove('file_example_XLSX_50.xlsx')


def test_csv():
    with ZipFile(zip_path) as arch:
        text = str(arch.read('users.csv'))
        assert text.__contains__('Mia,Levin,16/02/2001')


def test_pdf():
    with ZipFile(zip_path) as arch:
        arch.extract('Lapin_giftcard.pdf')
        text = PdfReader('Lapin_giftcard.pdf').pages[0].extract_text()
        assert text.__contains__('Открытка')
        os.remove('Lapin_giftcard.pdf')


def test_zip_deleted():
    os.remove(zip_path)
    assert len(os.listdir(files_path)) == 3



